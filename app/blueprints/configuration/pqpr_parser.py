"""
Parses the PQPR "FG -- Copy" sheet into a plain-JSON-friendly structure:

{
  "components": ["A50C", "A75C", ...],
  "kits": [
    {"edp": "0241276", "kit_name": "1675KIT48", "is_top10": true,
     "components": {"A75C": 1, "A50N": 1}},
    ...
  ]
}

All sheet/column locations are read from config.yaml (pqpr_parsing) rather
than hardcoded, since the sheet layout may change.
"""

import openpyxl
from openpyxl.utils import column_index_from_string


def parse_pqpr_workbook(filepath: str, parsing_config: dict) -> dict:
    sheet_name = parsing_config["sheet_name"]
    header_row = parsing_config["header_row"]
    kit_name_col = column_index_from_string(parsing_config["kit_name_column"])
    edp_col = column_index_from_string(parsing_config["edp_column"])
    component_start_col = column_index_from_string(
        parsing_config["component_start_column"]
    )
    top10_row_count = parsing_config["top10_row_count"]

    wb = openpyxl.load_workbook(filepath, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the uploaded file.")
    ws = wb[sheet_name]

    component_names = []
    for col in range(component_start_col, ws.max_column + 1):
        header_value = ws.cell(row=header_row, column=col).value
        if header_value is None or str(header_value).strip() == "":
            continue
        component_names.append((col, str(header_value).strip()))

    kits = []
    data_row_index = 0  # counts actual kit rows, used for top-10 ranking
    for row in range(header_row + 1, ws.max_row + 1):
        kit_name = ws.cell(row=row, column=kit_name_col).value
        edp = ws.cell(row=row, column=edp_col).value

        if kit_name is None or str(kit_name).strip() == "":
            continue

        kit_name = str(kit_name).strip()
        edp = str(edp).strip() if edp is not None else ""

        components = {}
        for col, comp_name in component_names:
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is None or str(cell_value).strip() == "":
                continue
            raw = str(cell_value).strip()
            if raw.lower() == "x":
                qty = 1
            else:
                try:
                    qty = int(float(raw))
                except ValueError:
                    continue
            if qty > 0:
                components[comp_name] = qty

        data_row_index += 1
        kits.append(
            {
                "edp": edp,
                "kit_name": kit_name,
                "is_top10": data_row_index <= top10_row_count,
                "components": components,
            }
        )

    return {
        "components": [name for _, name in component_names],
        "kits": kits,
    }
